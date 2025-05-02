from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
import mysql.connector
from functools import wraps
import pandas as pd
from flask import Response
from io import BytesIO
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Koneksi ke database
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="skhujpg9h2jx",
    database="document_v1"
)
cursor = mydb.cursor(dictionary=True)


# ======================== LOGIN SECTION ========================
@app.route('/')
def index():
    return redirect(url_for('login'))

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Silakan login terlebih dahulu.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        cursor.execute("SELECT * FROM users WHERE username = %s AND password = %s", (username, password))
        user = cursor.fetchone()

        if user:
            session['username'] = user['username']
            session['user_id'] = user['id']
            session['role'] = user['role']  # ⬅️ Simpan role di session
            return redirect(url_for('home'))

        else:
            flash('Username atau password salah!', 'danger')
            return redirect(url_for('login'))  # ⬅️ pastikan redirect ke login lagi jika gagal

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash("Berhasil logout!", "info")
    return redirect(url_for('login'))

# ======================== MAIN PAGE ========================
@app.route('/home')
@login_required
def home():
    if 'user_id' not in session:
        flash("Sesi login tidak valid. Silakan login ulang.", "warning")
        return redirect(url_for('login'))

    user_id = session['user_id']
    role = session.get('role')  # Ambil role dari session

    if role in ['Admin', 'Super Admin']:
        cursor.execute("SELECT * FROM project_event")
    else:
        cursor.execute("SELECT * FROM project_event WHERE user_id = %s", (user_id,))
    
    data = cursor.fetchall()

    # Jika role Admin, ambil nama pekerja untuk setiap project
    if role in ['Admin', 'Super Admin']:
        for row in data:
            cursor.execute("SELECT nama_pekerja FROM users WHERE id = %s", (row['user_id'],))
            pekerja = cursor.fetchone()
            row['user_by'] = pekerja['nama_pekerja'] if pekerja else 'Unknown'  # Menambahkan nama pekerja

    return render_template('home.html', username=session['username'], project_data=data)\
    

# ======================== DELETE EVENT ========================
@app.route('/delete_project/<int:project_id>', methods=['POST'])
@login_required
def delete_project(project_id):
    role = session.get('role')
    user_id = session.get('user_id')

    # Validasi hak akses
    if role in ['Admin', 'Super Admin']:
        cursor.execute("DELETE FROM project_event WHERE id = %s", (project_id,))
    else:
        # User biasa hanya boleh hapus miliknya sendiri
        cursor.execute("DELETE FROM project_event WHERE id = %s AND user_id = %s", (project_id, user_id))

    mydb.commit()
    # flash("Data berhasil dihapus.", "success")
    return redirect(url_for('home'))


# ======================== INPUT EVENT ========================
@app.route('/inputEvents', methods=['GET'])
@login_required
def inputEvents():
    return render_template('inputEvents.html')

@app.route('/submit-project', methods=['POST'])
@login_required
def submit_project():
    try:
        kode_project = request.form['kode_project']
        judul_project = request.form['judul_project']
        tanggal_event = request.form['tanggal_event']
        hari_kerja = request.form['hari_kerja']
        nama_venue = request.form['nama_venue']
        lokasi_venue = request.form['lokasi_venue']
        jenis_pekerjaan = request.form['jenis_pekerjaan']

        # Get multiple selected values from checkboxes
        waktu_pekerjaan_list = request.form.getlist('waktu_pekerjaan[]')
        waktu_pekerjaan = ', '.join(waktu_pekerjaan_list)  # Convert to comma-separated string

        keterangan = request.form['keterangan']
        approval_status = 'review'
        user_id = session['user_id']

        query = """
            INSERT INTO project_event (
                kode_project, judul_project, tanggal_event, hari_kerja,
                nama_venue, lokasi_venue, jenis_pekerjaan, waktu_pekerjaan, keterangan, approval_status, user_id
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        values = (
            kode_project, judul_project, tanggal_event, hari_kerja,
            nama_venue, lokasi_venue, jenis_pekerjaan, waktu_pekerjaan, keterangan, approval_status, user_id
        )

        cursor.execute(query, values)
        mydb.commit()

        flash("Data project berhasil ditambahkan!", "success")
        return redirect(url_for('inputEvents'))

    except Exception as e:
        flash(f"Terjadi kesalahan: {str(e)}", "danger")
        return redirect(url_for('inputEvents'))

# ======================== USER DATA ========================
@app.route('/dataUsers', methods=['GET'])
@login_required
def dataUsers():
    cursor.execute("SELECT * FROM users")
    users_data = cursor.fetchall()
    return render_template('dataUser.html', users=users_data)

@app.route('/delete_user/<string:username>', methods=['GET'])
@login_required
def delete_user(username):
    try:
        cursor.execute("DELETE FROM users WHERE username = %s", (username,))
        mydb.commit()
    except Exception as e:
        mydb.rollback()
        flash(f'Terjadi kesalahan saat menghapus user: {e}', 'danger')
    return redirect(url_for('dataUsers'))

@app.route('/registerUsers', methods=['GET', 'POST'])
@login_required
def registerUsers():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        no_pekerja = request.form['no_pekerja']
        nama_pekerja = request.form['nama_pekerja']
        role = request.form['role']

        cursor.execute("SELECT * FROM users WHERE no_pekerja = %s", (no_pekerja,))
        existing_user = cursor.fetchone()

        if existing_user:
            flash('Nomor Pekerja sudah terdaftar. Gunakan nomor lain.', 'danger')
            return redirect(url_for('registerUsers'))

        cursor.execute("""
            INSERT INTO users (username, password, no_pekerja, nama_pekerja, role) 
            VALUES (%s, %s, %s, %s, %s)
        """, (username, password, no_pekerja, nama_pekerja, role))
        mydb.commit()

        flash('User berhasil ditambahkan!', 'success')
        return redirect(url_for('registerUsers'))

    return render_template('registerUsers.html')

@app.route('/edit_user/<username>', methods=['GET', 'POST'])
@login_required
def edit_user(username):
    if request.method == 'POST':
        new_username = request.form['username']
        password = request.form['password']
        no_pekerja = request.form['no_pekerja']
        nama_pekerja = request.form['nama_pekerja']
        role = request.form['role']

        cursor.execute("""
            UPDATE users 
            SET username=%s, password=%s, no_pekerja=%s, nama_pekerja=%s, role=%s 
            WHERE username=%s
        """, (new_username, password, no_pekerja, nama_pekerja, role, username))
        mydb.commit()

        flash('User berhasil diperbarui!', 'success')
        return redirect(url_for('dataUsers'))

    cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
    user = cursor.fetchone()
    return render_template('editUser.html', user=user)
    
@app.route('/update-approval-status', methods=['POST'])
@login_required
def update_approval_status():
    try:
        data = request.get_json()
        project_id = data['project_id']
        approval_status = data['approval_status']

        # Perbarui status approval di database
        query = """
            UPDATE project_event
            SET approval_status = %s
            WHERE id = %s
        """
        values = (approval_status, project_id)

        cursor.execute(query, values)
        mydb.commit()

        return jsonify({'success': True})

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 400    

@app.route('/download_project_data')
@login_required
def download_project_data():
    if 'user_id' not in session:
        flash("Sesi login tidak valid. Silakan login ulang.", "warning")
        return redirect(url_for('login'))

    user_id = session['user_id']
    role = session.get('role')

    if role in ['Admin', 'Super Admin']:
        cursor.execute("SELECT * FROM project_event")
    else:
        cursor.execute("SELECT * FROM project_event WHERE user_id = %s", (user_id,))
    
    data = cursor.fetchall()

    # If the role is Admin, fetch the 'user_by' field (i.e., nama_pekerja)
    if role in ['Admin', 'Super Admin']:
        for row in data:
            cursor.execute("SELECT nama_pekerja FROM users WHERE id = %s", (row['user_id'],))
            pekerja = cursor.fetchone()
            row['user_by'] = pekerja['nama_pekerja'] if pekerja else 'Unknown'  # Adding user_by

    # Create a DataFrame from the query result
    columns = ['kode_project', 'judul_project', 'tanggal_event', 'hari_kerja', 'nama_venue', 'lokasi_venue', 'waktu_pekerjaan', 'jenis_pekerjaan', 'keterangan', 'approval_status', 'user_by', 'created_at']
    df = pd.DataFrame(data, columns=columns)

    # Create an Excel workbook and write the data to it
    wb = Workbook()
    ws = wb.active
    ws.title = 'Project Data'

    # Write the header row
    for col_num, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=col_name)

    # Write data rows
    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return the XLSX file as a response
    return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={"Content-Disposition": "attachment;filename=project_data.xlsx"})

@app.route('/edit_project/<kode_project>')
@login_required
def edit_event(kode_project):
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT * FROM project_event WHERE kode_project = %s", (kode_project,))
    project = cursor.fetchone()

    if not project:
        flash("Project tidak ditemukan.", "danger")
        return redirect(url_for('home'))

    return render_template("editEvent.html", project=project)

@app.route('/submit_project', methods=['POST'])
@login_required
def submit_event():
    kode = request.form['kode_project']
    judul = request.form['judul_project']
    tanggal = request.form['tanggal_event']
    hari = request.form['hari_kerja']
    venue = request.form['nama_venue']
    lokasi = request.form['lokasi_venue']

    waktu_list = request.form.getlist('waktu_pekerjaan[]')  # list of checked values
    waktu = ', '.join(waktu_list) if waktu_list else ''     # for multi-checkbox

    jenis = request.form.get('jenis_pekerjaan')  # single selection (dropdown)

    ket = request.form['keterangan']
    approval_status = request.form['approval_status']
    user_id = session['user_id']

    cursor.execute("SELECT * FROM project_event WHERE kode_project = %s", (kode,))
    existing = cursor.fetchone()

    if existing:
        # Update existing
        cursor.execute("""
            UPDATE project_event SET 
            judul_project=%s, tanggal_event=%s, hari_kerja=%s,
            nama_venue=%s, lokasi_venue=%s, waktu_pekerjaan=%s, jenis_pekerjaan=%s, 
            keterangan=%s, approval_status=%s
            WHERE kode_project=%s
        """, (judul, tanggal, hari, venue, lokasi, waktu, jenis, ket, approval_status, kode))
    else:
        # Insert new
        cursor.execute("""
            INSERT INTO project_event (
                kode_project, judul_project, tanggal_event, hari_kerja,
                nama_venue, lokasi_venue, waktu_pekerjaan, jenis_pekerjaan,
                keterangan, approval_status, user_id
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (kode, judul, tanggal, hari, venue, waktu, jenis, ket, approval_status, user_id))

    mydb.commit()
    return redirect(url_for('home'))


# ======================== ERROR HANDLER ========================
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

# ======================== RUN APP ========================
if __name__ == '__main__':
    app.run(debug=True)
